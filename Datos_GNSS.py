""" Sript for convert and join GNSS data
extract geodetic coordinate
Convert coordinate to CTM12 and 84
Generate GCP format
NV 2022
"""
import openpyxl.drawing.image
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


url=r"D:\MDA\Procesamiento GNSS\Seguimiento MDA.xlsx"

print("Input folder data: ",end="")
Dir=input()
print("specify the service number:",end="")
S=input();S=int(S)
csv0=Dir+"\\Puntos_Ajustados.csv" #Adjusted poinst REVISAR
xls=Dir+"\\GAM-FT-316 Ficha Puntos de control levantamientos Cartograficos.xlsx" #GCP Format
csv3=Dir+"\\PXYZ.csv" #Final points

dfIT = pd.read_excel(url)
# query frome records
qry = dfIT.query(f"ID=={S}")

df=pd.read_csv(csv0,encoding="latin-1")
df=df[["Name","X (m)","Y (m)","Z (m)","N Standard Deviation (m)","E Standard Deviation (m)","U Standard Deviation (m)"]]
df.columns=["ID","X","Y","Z","sdY","sdX","sdZ"]
df0=df
#Extract Geodetic coordinate ----------------------------------------------------------------------------------
print("Input tracking date (dd/mm/aaaa):",end="")
Fecha_Rast=input()
dfGeodetic=df[["ID","X","Y","Z"]]
dfGeodetic=dfGeodetic.assign(Fecha_Rast=Fecha_Rast)
dfGeodetic=dfGeodetic.assign(Fecha_Ref="01/01/2018")
dfGeodetic.to_csv(Dir+"\\XYZ.csv",index=False)
#break
print("Please project the data and press any key:",end="")
input()
#Read the other df
csv1=Dir+"\\84.csv"; csv2=Dir+"\\CTM12.csv"
df1=pd.read_csv(csv1); df2=pd.read_csv(csv2)
df3=pd.merge(df1,df2)
df3=df3[["ID","Latitud","Longitud","Norte","Este","Altura","Ondulacion"]]
df3["Altura_Orto"]=(df3["Altura"]-df3["Ondulacion"])
df3=pd.merge(df3,df)
df3=df3.drop(columns=["X","Y","Z"])

""" """
#Write the xls file
df=df3
#Get the xls file
wb=load_workbook(xls)
#wb sheets
sheets=wb.sheetnames
for sheet in sheets:
    if sheet != sheets[0]:
        del wb[sheet]

for point in df.index:
    #Variables
    id=(df["ID"][point]);lat = (df["Latitud"][point]);lon = (df["Longitud"][point]);N = (df["Norte"][point]);E = (df["Este"][point]);Z = (df["Altura"][point]);Zort = (df["Altura_Orto"][point]);sdY = (df["sdY"][point]);sdX = (df["sdX"][point]);sdZ = (df["sdZ"][point])

    #Duplicate sheets
    try:
        hj=wb[str(id)]
    except:
        pass
        hjdest=wb.copy_worksheet(hj)
        hjdest.title=str(id)
        # Add image DLIA
        img = openpyxl.drawing.image.Image(r"D:\MDA\DLIA.png")
        img.anchor = "A1"
        img.height = 100
        img.width = 170
        hjdest.add_image(img)
    #statics variables
    Nptos=len(df)
    #Write static data from google sheets

    Dirs = [("DLIA", "Dirección de Laboratorio e Innovación Ambiental"), ("DGEN", "Dirección General"),
            ("OTIC", "Oficina de las Tecnologías de la Información y las Comunicaciones"),
            ("FIAB", "Fondo para las Inversiones Ambientales en la Cuenca del Río Bogotá"),
            ("DRN", "Dirección de Recursos Naturales"),
            ("DGOAT", "Dirección de Gestión del Ordenamiento Ambiental y Territorial"),
            ("DESCA", "Dirección de Evaluación, Seguimiento y Control Ambiental"), ("DJUR", "Dirección Jurídica"),
            ("DIA", "Dirección de Infraestructura Ambiental"),
            ("DCASC", "Dirección de Cultura Ambiental y Servicio al Ciudadano"),
            ("DRBC", "Dirección Regional Bogotá- La Calera"), ("DRAG", "Dirección Regional Almeidas y Guatavita"),
            ("DRAM", "Dirección Regional Alto Magdalena"), ("DRBM", "Dirección Regional Bajo Magdalena"),
            ("DRCH", "Dirección Regional Chiquinquirá"), ("DRGU", "Dirección Regional Gualivá"),
            ("DRMC", "Dirección Regional Magdalena Centro"), ("DRRN", "Dirección Regional Rionegro"),
            ("DRSC", "Dirección Regional Sabana Centro"), ("DRSO", "Dirección Regional Sabana Occidente"),
            ("DRSOA", "Dirección Regional Soacha"), ("DRSU", "Dirección Regional Sumapáz"),
            ("DRTE", "Dirección Regional Tequendama"),
            ("DRUB", "Dirección Regional Ubaté")]  # tupla Direcciónes Generales

    Solicitante = qry.iloc[0]["Solicitante"]
    for item in Dirs:
        if item[0] == Solicitante:
            Dirección_Car = item[1]
            break

    proyecto = qry.iloc[0]["Título"]
    municipio = qry.iloc[0]['Municipio']
    Depa = [("Buenavista", "Boyacá"), ("Caldas", "Boyacá"), ("Chiquinquierá", "Boyacá"), ("Ráquira", "Boyacá"),
            ("Saboyá", "Boyacá"), ("San miguel de sema", "Boyacá")]
    for item1 in Depa:
        if item1[0] == municipio:
            departamento = item1[1]
        else:
            departamento = "Cundinamarca"
    piloto = qry.iloc[0]["Piloto"]
    observador = qry.iloc[0]["Observador"]
    drone = qry.iloc[0]["Equipo"]
    fecha = qry.iloc[0]["Fecha Fin Vuelos"]

    sheet=wb[str(id)]
    sheet["B3"] = Solicitante  # Solicitante
    sheet["E3"] = proyecto  # Proyecto
    sheet["B4"] = departamento  # Departamento
    sheet["E4"] = piloto  # Piloto
    sheet["B5"] = municipio  # Municipio
    sheet["E5"] = observador  # Observador
    sheet["B8"] = drone  # drone
    sheet["F7"] = id
    sheet["F8"] = Nptos
    sheet["F9"] = 2  # AltIns
    sheet["F10"] = fecha  # Fecha

    #write dynaimc variables
    sheet["A13"] = lat
    sheet["B13"]=lon
    sheet["c13"] =Z
    sheet["D13"] =N
    sheet["E13"] =E
    sheet["F13"] =Zort
    sheet["D15"] =sdX
    sheet["E15"] =sdY
    sheet["F15"] =sdZ
sheets=wb.sheetnames

#write all data
sheet_data=wb.create_sheet()
sheet_data.title="Datos"
ws=wb["Datos"]
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)
for cell in ws['A'] + ws[1]:
    cell.style = 'Pandas'
wb.save(xls)
#Export final points
df4=df3[["ID","Este","Norte","Altura_Orto"]]
df4.to_csv(csv3,index=False)
print("GAM-FT-316 Sucessfull!!")


